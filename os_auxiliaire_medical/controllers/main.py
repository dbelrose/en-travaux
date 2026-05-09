from odoo import http
from odoo.http import request, Response
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class CpsBordereauController(http.Controller):

    @http.route('/cps/bordereau/<int:bordereau_id>/xlsx', type='http', auth='user')
    def export_bordereau_xlsx(self, bordereau_id, **kwargs):
        bordereau = request.env['cps.bordereau'].browse(bordereau_id)
        if not bordereau.exists():
            return Response('Bordereau introuvable', status=404)

        wb = Workbook()
        ws = wb.active
        ws.title = "Bordereau"

        GREEN = "1F6B3A"
        LIGHT_GREEN = "E8F5E9"

        def thin():
            s = Side(style='thin', color='AAAAAA')
            return Border(left=s, right=s, top=s, bottom=s)

        for col, w in zip("ABCDEFG", [6, 28, 14, 13, 13, 16, 18]):
            ws.column_dimensions[col].width = w

        row = 1
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = "BORDEREAU DE FACTURATION"
        ws[f"A{row}"].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=GREEN)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 1

        for i, (lbl, val) in enumerate([
            ("N° bordereau :", bordereau.name),
            ("Date :", bordereau.date_bordereau.strftime('%d.%m.%Y') if bordereau.date_bordereau else ''),
            ("Nb factures :", str(bordereau.nb_feuilles)),
        ]):
            c1, c2 = ["A", "D", "F"][i], ["B", "E", "G"][i]
            ws[f"{c1}{row}"] = lbl
            ws[f"{c1}{row}"].font = Font(bold=True, size=10)
            ws[f"{c2}{row}"] = val
            ws[f"{c2}{row}"].font = Font(size=10)
        row += 1

        prat = bordereau.praticien_id
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = f"{prat.vat or ''}  ·  {prat.name}  ·  Tél : {prat.phone or ''}  ·  {prat.street or ''}"
        ws[f"A{row}"].font = Font(size=10, italic=True, color=GREEN)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = bordereau.mois.upper()
        ws[f"A{row}"].font = Font(bold=True, size=11, color=GREEN)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 2

        headers = ["N°", "Nom Prénom", "DN", "Soins DU", "Soins AU", "Pmt CPS", "Pmt Patient"]
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=i + 1, value=h)
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=GREEN)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin()
        ws.row_dimensions[row].height = 22
        row += 1

        first_data = row
        lignes = bordereau.get_lignes_for_report()
        for idx, l in enumerate(lignes):
            bg = "F5F5F5" if idx % 2 == 0 else "FFFFFF"
            data = [l['n'], l['nom_prenom'], l['dn'], l['date_debut'], l['date_fin'],
                    l['montant_cps'], l['montant_patient']]
            for ci, val in enumerate(data):
                c = ws.cell(row=row, column=ci + 1, value=val)
                c.border = thin()
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center" if ci != 1 else "left", vertical="center")
                c.font = Font(size=10)
                if ci in (5, 6):
                    c.number_format = '#,##0 "F"'
            row += 1

        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=GREEN)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"A{row}"].border = thin()
        for col_letter in ["F", "G"]:
            c = ws[f"{col_letter}{row}"]
            c.value = f"=SUM({col_letter}{first_data}:{col_letter}{row - 1})"
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=GREEN)
            c.alignment = Alignment(horizontal="center")
            c.border = thin()
            c.number_format = '#,##0 "F"'

        ws.freeze_panes = "A7"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Bordereau_{bordereau.name}_{bordereau.mois.replace(' ', '_')}.xlsx"
        return Response(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=[('Content-Disposition', f'attachment; filename="{filename}"')]
        )
